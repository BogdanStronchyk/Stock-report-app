import os
import re
from typing import Any, Dict, Optional, Tuple
from openpyxl import load_workbook

_NUM = r"[+-]?\d*\.?\d+"  # signed float

def _norm_metric(s: str) -> str:
    s = s or ""
    s = s.strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s

def _strip_parens(s: str) -> str:
    # remove parenthetical fragments to improve matching: "FCF Yield (TTM ...)" -> "FCF Yield"
    return re.sub(r"\s*\([^)]*\)", "", s).strip()

def _is_heading_row(metric: str) -> bool:
    m = _norm_metric(metric)
    if not m:
        return True
    # common non-metric helper rows in the Sector Adjustments sheet
    return any(
        x in m
        for x in [
            "how to use",
            "step",
            "sector id",
            "adjustments",
            "notes",
        ]
    )

def _token_set(s: str) -> set:
    s = _norm_metric(_strip_parens(s))
    # split on punctuation too
    s = re.sub(r"[^a-z0-9%/\- ]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return set([t for t in s.split(" ") if t])

def _metric_matches(adj_metric: str, threshold_metric: str) -> bool:
    """Match a Sector Adjustments row metric to a category-sheet metric name.

    Rules are intentionally conservative to prevent collisions like:
      "Market Cap" accidentally matching "SBC % of Market Cap (TTM)".
    """
    a_raw = adj_metric or ""
    t_raw = threshold_metric or ""

    a = _norm_metric(a_raw)
    t = _norm_metric(t_raw)
    if not a or not t:
        return False

    # 1) Exact match
    if a == t:
        return True

    # 2) Exact after stripping parentheses on either side
    a2 = _norm_metric(_strip_parens(a_raw))
    t2 = _norm_metric(_strip_parens(t_raw))
    if a2 and a2 == t2:
        return True

    # 3) Strong safeguards for common collisions
    # If one mentions SBC and the other doesn't -> no match
    if ("sbc" in a2) != ("sbc" in t2):
        return False
    # If one has a percent sign and the other doesn't -> no match
    if ("%" in a2) != ("%" in t2):
        return False
    # If one has 'market cap' but not the other -> no match
    if ("market cap" in a2) != ("market cap" in t2):
        return False

    # 4) Allow safe prefix / containment only for a small whitelist
    safe_prefixes = (
        "fcf yield",
        "ev/fcf",
        "margin trend",
        "roic",
        "share count cagr",
        "net buyback yield",
        "shareholder yield",
        "short interest",
        "days to cover",
        "avg daily",
        "max drawdown",
        "realized volatility",
        "worst weekly return",
        "p/e",
        "p/s",
        "ev/ebit",
        "ev/ebitda",
    )
    if any(a2.startswith(sp) for sp in safe_prefixes) or any(t2.startswith(sp) for sp in safe_prefixes):
        if a2 and (a2 in t2 or t2 in a2):
            return True

    # 5) Token overlap similarity (conservative)
    ta = _token_set(a_raw)
    tt = _token_set(t_raw)
    if not ta or not tt:
        return False

    overlap = len(ta & tt) / max(len(ta), len(tt))
    # require very high overlap and similar length
    if overlap >= 0.85 and abs(len(a2) - len(t2)) <= 12:
        return True

    return False


def parse_range_cell(s: Any) -> Optional[Tuple[Optional[float], Optional[float]]]:
    """Robustly parse threshold text into (lo, hi).

    Handles strings like:
      "< 15"
      "15–25"
      "> 25 or negative"
      "> 6% | otherwise ..."
      "< 3 days"
      "$10B" / "$250M" / "$1.2T"
      "-35 to -50"

    Returns (None, hi) for "< hi", (lo, None) for "> lo", (lo, hi) for ranges.
    """
    if s is None:
        return None
    txt = str(s).strip()
    if not txt:
        return None

    # Normalize dashes and whitespace
    txt = txt.replace("–", "-").replace("—", "-")
    txt = re.sub(r"\s+", " ", txt).strip()

    # Ignore purely qualitative guidance
    low_txt = txt.lower()
    if any(w in low_txt for w in ["expanding", "stable", "contracting"]):
        return None

    # Remove separators that interfere with float parsing
    txt = txt.replace(",", "").replace("_", "")

    # Currency multipliers (only if $ is present)
    mul = 1.0
    if "$" in txt:
        m = re.search(rf"({_NUM})\s*([KMBT])\b", txt, flags=re.IGNORECASE)
        if m:
            suf = m.group(2).upper()
            mul = {"K": 1e3, "M": 1e6, "B": 1e9, "T": 1e12}.get(suf, 1.0)
            txt = re.sub(rf"({_NUM})\s*[KMBT]\b", r"\1", txt, flags=re.IGNORECASE)
        txt = txt.replace("$", "").strip()

    # Remove % (metrics are stored as percent points, not fractions)
    txt = txt.replace("%", "").strip()

    # Normalize "to" to "-"
    txt = re.sub(r"\s+to\s+", " - ", txt, flags=re.IGNORECASE)

    # 1) Comparator rule
    m = re.search(rf"(<=|>=|<|>)\s*({_NUM})", txt)
    if m:
        op = m.group(1)
        num = float(m.group(2)) * mul
        if op in ("<", "<="):
            return (None, num)
        else:
            return (num, None)

    # 2) Range rule
    m = re.search(rf"({_NUM})\s*-\s*({_NUM})", txt)
    if m:
        a = float(m.group(1)) * mul
        b = float(m.group(2)) * mul
        lo, hi = (a, b) if a <= b else (b, a)
        return (lo, hi)

    return None


def load_thresholds_from_excel(path: str) -> Dict[str, Dict[str, Dict[str, Any]]]:
    if not os.path.exists(path):
        raise FileNotFoundError(f"Checklist file not found: {path}")

    wb = load_workbook(path, data_only=True)

    categories = ["Valuation", "Profitability", "Balance Sheet", "Growth", "Risk"]
    thresholds: Dict[str, Dict[str, Dict[str, Any]]] = {c: {} for c in categories}

    for cat in categories:
        if cat not in wb.sheetnames:
            continue
        ws = wb[cat]
        for r in range(2, ws.max_row + 1):
            metric = ws.cell(r, 1).value
            if not metric or not isinstance(metric, str):
                continue

            green = ws.cell(r, 2).value
            yellow = ws.cell(r, 3).value
            red = ws.cell(r, 4).value
            marking = ws.cell(r, 5).value
            notes = ws.cell(r, 6).value

            thresholds[cat][metric.strip()] = {
                "Default (All)": {
                    "green_txt": green,
                    "yellow_txt": yellow,
                    "red_txt": red,
                    "marking": marking,
                    "notes": notes,
                }
            }

    # Sector adjustments sheet (optional)
    if "Sector Adjustments" in wb.sheetnames:
        ws = wb["Sector Adjustments"]
        header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        col_map = {name: idx + 1 for idx, name in enumerate(header) if isinstance(name, str)}

        sector_cols = [
            "Default (All)",
            "Software/Tech",
            "Industrials",
            "Consumer Staples",
            "Consumer Discretionary",
            "Healthcare/Pharma",
            "Energy/Materials",
            "Financials (Banks)",
            "REITs",
            "Utilities/Telecom",
        ]

        notes_col = col_map.get("Notes (why/when)", ws.max_column)

        for r in range(2, ws.max_row + 1):
            metric = ws.cell(r, col_map.get("Metric", 1)).value
            if not metric or not isinstance(metric, str):
                continue
            metric = metric.strip()
            if _is_heading_row(metric):
                continue

            for sec in sector_cols:
                if sec not in col_map:
                    continue
                val = ws.cell(r, col_map[sec]).value
                if val is None or not isinstance(val, str) or not val.strip():
                    continue

                parts = [p.strip() for p in val.split("/") if p.strip()]
                if len(parts) < 3:
                    continue

                green_txt, yellow_txt, red_txt = parts[0], parts[1], parts[2]
                sec_notes = ws.cell(r, notes_col).value

                for cat in thresholds:
                    for tm in list(thresholds[cat].keys()):
                        if _metric_matches(metric, tm):
                            thresholds[cat][tm][sec] = {
                                "green_txt": green_txt,
                                "yellow_txt": yellow_txt,
                                "red_txt": red_txt,
                                "marking": thresholds[cat][tm]["Default (All)"].get("marking"),
                                "notes": sec_notes,
                            }

    return thresholds


def get_threshold_set(
    thresholds: Dict[str, Dict[str, Dict[str, Any]]],
    category: str,
    metric: str,
    sector_bucket: str
) -> Dict[str, Any]:
    if category not in thresholds or metric not in thresholds[category]:
        return {}
    by_sector = thresholds[category][metric]
    if sector_bucket in by_sector:
        return by_sector[sector_bucket]
    return by_sector.get("Default (All)", {})
