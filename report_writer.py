from typing import Any, Dict, List, Optional, Tuple
import math

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

from config import FILL_HDR, FONT_HDR, ALIGN_CENTER, ALIGN_WRAP, FILL_GREEN, FILL_YELLOW, FILL_RED, FILL_GRAY
from checklist_loader import get_threshold_set, parse_range_cell
from scoring import score_with_threshold_txt, compute_category_score_and_coverage, adjusted_from_raw_and_coverage
from eligibility import evaluate_eligibility

# --- CORE 12 WHITELIST ---
WHITELIST = {
    "P/E (TTM, positive EPS)", "EV/EBIT", "FCF Yield (TTM FCF / Market Cap)",
    "Gross Margin %", "Operating Margin %", "ROIC % (standardized)",
    "Net Debt / EBITDA", "Interest Coverage (EBIT / Interest)",
    "Revenue per Share CAGR (5Y)", "FCF per Share CAGR (5Y)",
    "Market Cap", "Max Drawdown (3–5Y)"
}


# --- HELPERS ---
def _is_percent_like_metric(metric: str) -> bool:
    m = metric.lower()
    return "%" in metric or "yield" in m or "margin" in m or "drawdown" in m or "cagr" in m or "roic" in m


def _apply_metric_value_format(cell, metric: str, val: Any):
    if val is None or not isinstance(val, (int, float)): return
    if _is_percent_like_metric(metric):
        cell.number_format = '0.00"%"'
    else:
        cell.number_format = "0.00"


def band_fill(score: Optional[float]):
    if score is None: return FILL_GRAY
    try:
        s = float(score)
    except:
        return FILL_GRAY
    if s >= 60.0: return FILL_GREEN
    if s >= 40.0: return FILL_YELLOW
    return FILL_RED


def reversal_fill(score: Optional[float]):
    if score is None: return FILL_GRAY
    try:
        s = float(score)
    except:
        return FILL_GRAY
    if s >= 70.0: return FILL_GREEN
    if s >= 50.0: return FILL_YELLOW
    return FILL_RED


def _limits_text(th: Optional[Dict[str, Any]]) -> str:
    if not th: return ""
    parts = []
    for k, p in [("green_txt", "G"), ("yellow_txt", "Y"), ("red_txt", "R")]:
        v = th.get(k)
        if v: parts.append(f"{p}:{v}")
    return " | ".join(parts)


def _metric_weight(metric: str) -> float:
    if metric == "EV/EBIT" or metric == "FCF Yield (TTM FCF / Market Cap)": return 2.0
    if metric == "ROIC % (standardized)": return 2.0
    if metric == "Net Debt / EBITDA": return 1.5
    return 1.0


def final_recommendation_banner(fund_adj: Optional[float], reversal_total: Optional[float]) -> Tuple[str, Any]:
    f = fund_adj if fund_adj is not None else 0
    r = reversal_total if reversal_total is not None else 0

    if f >= 60 and r >= 60: return ("✅ STRONG BUY — High Quality + Trend", FILL_GREEN)
    if f >= 60 and r < 50: return ("⚠ WATCH — Quality is good, wait for entry", FILL_YELLOW)
    if f < 40: return ("❌ AVOID — Fundamental risks", FILL_RED)
    if r >= 70 and f < 40: return ("⛔ TRAP — Momentum without Quality", FILL_RED)
    return ("⚠ HOLD / NEUTRAL", FILL_YELLOW)


def _normalize_reversal_pack(revpack: Dict[str, Any]) -> Dict[str, Any]:
    if not isinstance(revpack, dict): revpack = {}
    fs = revpack.get("fund_score_pct") or revpack.get("fundamental_score")
    ts = revpack.get("tech_score_pct") or revpack.get("technical_score")
    total = revpack.get("total_score_pct")

    if total is None and fs is not None and ts is not None:
        try:
            total = 0.6 * float(fs) + 0.4 * float(ts)
        except:
            total = None

    return {
        "fund_symbols": revpack.get("fund_symbols", {}) or revpack.get("fundamental_symbols", {}),
        "tech_symbols": revpack.get("tech_symbols", {}) or revpack.get("technical_symbols", {}),
        "fund_details": revpack.get("fund_details", {}) or revpack.get("fundamental", {}),
        "tech_details": revpack.get("tech_details", {}) or revpack.get("technical", {}),
        "fund_score_pct": fs,
        "tech_score_pct": ts,
        "total_score_pct": total,
    }


def _write_reversal_block(ws, start_row: int, title: str, symbols: Dict[str, str], details: Dict[str, Tuple[int, str]],
                          score: Optional[float]) -> int:
    ws[f"A{start_row}"] = title
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=6)
    start_row += 1

    headers = ["Condition", "Score", "Details"]
    for i, h in enumerate(headers):
        c = ws.cell(start_row, i + 1);
        c.value = h;
        c.fill = FILL_HDR;
        c.font = FONT_HDR
    ws.cell(start_row, 4).value = f"Score: {float(score):.1f}%" if score is not None else "Score: NA"
    start_row += 1

    for cond, sym in symbols.items():
        ws.cell(start_row, 1).value = cond
        ws.cell(start_row, 2).value = sym
        det = details.get(cond)
        det_text = det[1] if isinstance(det, (list, tuple)) and len(det) > 1 else str(det)
        ws.cell(start_row, 3).value = det_text
        start_row += 1
    return start_row + 1


def autosize_columns(ws):
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)  # Safer than .column_letter for merged cells
        for cell in col:
            try:
                if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
            except:
                pass

        adjusted_width = min(max_length + 2, 60)
        adjusted_width = max(adjusted_width, 10)
        ws.column_dimensions[column].width = adjusted_width


def _add_cheat_sheet(wb: Workbook):
    ws = wb.create_sheet("Cheat Sheet", 1)

    # Title
    ws["A1"] = "HOW TO READ THIS REPORT"
    ws.merge_cells("A1:C1")
    ws["A1"].font = Font(size=14, bold=True)

    # 1. Scoring Bands
    ws["A3"] = "1. SCORING BANDS"
    ws["A3"].font = Font(bold=True)
    ws["A4"] = "GREEN";
    ws["B4"] = "≥ 60%";
    ws["C4"] = "Strong / High Quality"
    ws["A4"].fill = FILL_GREEN
    ws["A5"] = "YELLOW";
    ws["B5"] = "40 – 59%";
    ws["C5"] = "Acceptable / Average"
    ws["A5"].fill = FILL_YELLOW
    ws["A6"] = "RED";
    ws["B6"] = "< 40%";
    ws["C6"] = "Weak / Risky"
    ws["A6"].fill = FILL_RED

    # 2. Recommendations
    ws["A8"] = "2. RECOMMENDATION LOGIC"
    ws["A8"].font = Font(bold=True)
    ws["A9"] = "✅ STRONG BUY";
    ws["B9"] = "Fund Score ≥ 60  AND  Reversal Score ≥ 60";
    ws["C9"] = "High Quality + Trend Confirmation"
    ws["A10"] = "⚠ WATCH";
    ws["B10"] = "Fund Score ≥ 60  (but weak trend)";
    ws["C10"] = "Quality is good, wait for entry"
    ws["A11"] = "⛔ TRAP";
    ws["B11"] = "Reversal Score ≥ 70  (but weak fund)";
    ws["C11"] = "Momentum without Quality (Dangerous)"
    ws["A12"] = "❌ AVOID";
    ws["B12"] = "Fund Score < 40";
    ws["C12"] = "Fundamental risks too high"

    # 3. Core Metrics Explanation
    ws["A14"] = "3. CORE 12 METRICS EXPLAINED"
    ws["A14"].font = Font(bold=True)

    metrics_info = [
        ("Valuation", "EV/EBIT", "Enterprise Value / Earnings Before Interest & Tax. The purest valuation metric."),
        ("Valuation", "FCF Yield", "Free Cash Flow / Market Cap. The cash return you get on your purchase."),
        ("Valuation", "P/E", "Price / Earnings. Standard metric, used here for reference."),
        ("Quality", "ROIC %", "Return on Invested Capital. Measures efficiency and moat."),
        ("Quality", "Gross Margin", "% of Revenue left after COGS. Proxy for pricing power."),
        ("Quality", "Op Margin", "% of Revenue left after operating costs. Measures operational efficiency."),
        ("Safety", "Net Debt/EBITDA", "Years to pay off debt using earnings. > 3.0x is risky."),
        ("Safety", "Interest Cov", "EBIT / Interest Expense. Can they pay their bills? < 3.0x is risky."),
        ("Growth", "Rev/Share CAGR", "Revenue Growth per Share. The only 'top line' growth that counts."),
        ("Growth", "FCF/Share CAGR", "Cash Flow Growth per Share. The 'bottom line' growth reality."),
        ("Risk", "Drawdown", "Maximum drop from highs in 3-5 years. Measures volatility risk."),
        ("Risk", "Market Cap", "Size of company. Smaller caps often have higher liquidity risk.")
    ]

    row = 15
    ws["A15"] = "Category";
    ws["B15"] = "Metric";
    ws["C15"] = "What it tells you"
    for c in ["A15", "B15", "C15"]: ws[c].font = Font(bold=True); ws[c].fill = FILL_HDR

    for cat, name, desc in metrics_info:
        row += 1
        ws[f"A{row}"] = cat
        ws[f"B{row}"] = name
        ws[f"C{row}"] = desc

    autosize_columns(ws)


# ==========================================
# MAIN REPORT GENERATOR
# ==========================================
def create_report_workbook(tickers: List[str], thresholds: Dict[str, Dict[str, Dict[str, Any]]],
                           metrics_by_ticker: Dict[str, Dict[str, Any]], reversal_by_ticker: Dict[str, Dict[str, Any]],
                           out_path: str):
    wb = Workbook();
    wb.remove(wb.active)

    # --- 1. SUMMARY SHEET ---
    ws_sum = wb.create_sheet("Summary", 0)
    sum_headers = ["Ticker", "Sector", "Fund Score", "Reversal Score", "Recommendation",
                   "Valuation", "Quality", "Safety", "Growth", "Risk", "Coverage %"]
    ws_sum.append(sum_headers)
    for c in ws_sum[1]: c.fill = FILL_HDR; c.font = FONT_HDR; c.alignment = ALIGN_CENTER

    # --- 2. CHEAT SHEET ---
    _add_cheat_sheet(wb)

    category_maps = {"Valuation": "Valuation", "Profitability": "Quality", "Balance Sheet": "Safety",
                     "Growth": "Growth", "Risk": "Risk"}

    for t in tickers:
        m = metrics_by_ticker.get(t, {})
        revpack = _normalize_reversal_pack(reversal_by_ticker.get(t, {}))
        bucket = m.get("Sector Bucket", "Default (All)")

        ws = wb.create_sheet(t)

        # --- TITLE BLOCK ---
        ws["A1"] = f"{t} — {bucket}"
        ws.merge_cells("A1:C1")
        ws["A1"].font = Font(size=14, bold=True)

        # --- DETAILED TABLES ---
        cat_scores = {}
        cat_coverages = {}
        row = 3

        for cat_sheet, cat_display in category_maps.items():
            ws.cell(row, 1).value = cat_display
            ws.cell(row, 1).font = Font(bold=True, size=12)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            row += 1

            headers = ["Metric", "Value", "Rating", "Mode", "Limits", "Notes"]
            for i, h in enumerate(headers):
                c = ws.cell(row, i + 1);
                c.value = h;
                c.fill = FILL_HDR;
                c.font = FONT_HDR
            row += 1

            ratings = {}
            weights = {}

            th_set = thresholds.get(cat_sheet, {})
            for metric, limits in th_set.items():
                if metric not in WHITELIST: continue

                val = m.get(metric)
                th = get_threshold_set(thresholds, cat_sheet, metric, bucket)

                rating = "NA";
                fill = FILL_GRAY
                if th and val is not None:
                    # SCORING UNPACK (2 values)
                    rating, fill = score_with_threshold_txt(val, th.get("green_txt"), th.get("yellow_txt"),
                                                            th.get("red_txt"))

                ratings[metric] = rating
                weights[metric] = _metric_weight(metric)

                ws.cell(row, 1).value = metric
                c = ws.cell(row, 2, val);
                _apply_metric_value_format(c, metric, val)
                ws.cell(row, 3, rating).fill = fill
                ws.cell(row, 4).value = bucket
                ws.cell(row, 5).value = _limits_text(th)
                ws.cell(row, 6).value = th.get("notes", "")
                row += 1

            raw, cov = compute_category_score_and_coverage(ratings, weights)
            adj_score = adjusted_from_raw_and_coverage(raw, cov)
            cat_scores[cat_display] = adj_score
            cat_coverages[cat_display] = cov

            ws.cell(row, 1).value = f"{cat_display} Adjusted Score: {adj_score:.1f}%" if adj_score is not None else "NA"
            ws.cell(row, 1).font = Font(bold=True)
            row += 2

        # --- AGGREGATE SCORES ---
        valid_cats = [s for s in cat_scores.values() if s is not None]
        fund_score = sum(valid_cats) / len(valid_cats) if valid_cats else 0.0

        rev_total = revpack.get("total_score_pct")
        rec_txt, rec_fill = final_recommendation_banner(fund_score, rev_total)

        # --- DASHBOARD ---
        ws["D1"] = "Fundamental Score";
        ws["E1"] = fund_score;
        ws["E1"].fill = band_fill(fund_score)
        ws["D2"] = "Reversal Score";
        ws["E2"] = rev_total;
        ws["E2"].fill = reversal_fill(rev_total)
        ws["A2"] = "Recommendation";
        ws["B2"] = rec_txt;
        ws["B2"].fill = rec_fill

        # Summary Row
        avg_cov = sum(cat_coverages.values()) / len(cat_coverages) if cat_coverages else 0
        ws_sum.append([
            t, bucket, fund_score, rev_total, rec_txt,
            cat_scores.get("Valuation"), cat_scores.get("Quality"), cat_scores.get("Safety"),
            cat_scores.get("Growth"), cat_scores.get("Risk"), avg_cov
        ])

        # Reversal Blocks
        row = _write_reversal_block(ws, row, "Fundamental Turnaround", revpack.get("fund_symbols", {}),
                                    revpack.get("fund_details", {}), revpack.get("fund_score_pct"))
        row = _write_reversal_block(ws, row, "Technical Confirmation", revpack.get("tech_symbols", {}),
                                    revpack.get("tech_details", {}), revpack.get("tech_score_pct"))

        autosize_columns(ws)

    autosize_columns(ws_sum)
    wb.save(out_path)