from openpyxl.styles import PatternFill, Font, Alignment
import os

CHECKLIST_FILE = "Fundamental_Checklist_v3_value_matrix_fixed.xlsx"

# If True, always try Financial Modeling Prep (FMP) as a fallback when Yahoo/yfinance
# data is missing or incomplete.
FORCE_FMP_FALLBACK = (os.environ.get("FORCE_FMP_FALLBACK","").strip().lower() in ("1","true","yes"))
FILL_GREEN  = PatternFill("solid", fgColor="C6EFCE")
FILL_YELLOW = PatternFill("solid", fgColor="FFEB9C")
FILL_RED    = PatternFill("solid", fgColor="FFC7CE")
FILL_GRAY   = PatternFill("solid", fgColor="E7E6E6")
FILL_HDR    = PatternFill("solid", fgColor="1F4E79")

FONT_HDR = Font(color="FFFFFF", bold=True)
FONT_BOLD = Font(bold=True)

ALIGN_WRAP = Alignment(wrap_text=True, vertical="top")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

NUPL_REGIMES = [
    (0.6, float("inf"), "Euphoria"),
    (0.2, 0.6, "Optimism"),
    (-0.2, 0.2, "Neutral"),
    (-0.4, -0.2, "Stress"),
    (-0.6, -0.4, "Capitulation"),
    (float("-inf"), -0.6, "Deep Capitulation"),
]